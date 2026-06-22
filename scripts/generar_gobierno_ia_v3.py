"""
Genera la plantilla Gobierno IA V3 completa para Calzatura Vilchez.
Fuente: bolts empresariales, procesos de trazabilidad, módulo IA (v1.2) y corpus Q1.

Versión plantilla (estructura): V3
Versión contenido Calzatura Vilchez: ver CONTENT_VERSION

Uso: python scripts/generar_gobierno_ia_v3.py
"""
from __future__ import annotations

import shutil
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT_ARTIFACTS = ROOT / "artifacts" / "matrices" / "Plantilla_Gobierno_IA_V3_CALZATURA_VILCHEZ.xlsx"
TEMPLATE_DOWNLOADS = Path.home() / "Downloads" / "Plantilla_Gobierno_IA_V3.xlsx"
TEMPLATE = TEMPLATE_DOWNLOADS if TEMPLATE_DOWNLOADS.exists() else OUT_ARTIFACTS
BOLT_MATRIX = ROOT / "artifacts" / "matrices" / "Matriz_Registro_Web_Empresarial_AI_DLC_BOLT.xlsx"
OUT_DOWNLOADS = Path.home() / "Downloads" / "Plantilla_Gobierno_IA_V3_CALZATURA_VILCHEZ.xlsx"

PLANTILLA_VERSION = "V3"
CONTENT_VERSION = "1.2"
WEB_APP_VERSION = "1.9.0"
TODAY = date.today().isoformat()
OWNER = "Serpa Sedano Yeferson Wilson"
AGENT = "Cursor Agent / revisión técnica automatizada"
BRANCH = "main"

# Línea de releases del sistema web (calzatura-vilchez) — sincronizada con migraciones y git
VERSIONES_WEB = [
    ("1.0.0", "2026-05-01", "Bootstrap", "Catálogo, auth Firebase, pedidos base, Supabase create_all_tables.", "MP-01..06"),
    ("1.1.0", "2026-05-14", "Stock y ventas", "movimientos_stock, RPC registrar ingreso, ventas admin atómicas.", "MP-08, MP-11"),
    ("1.2.0", "2026-05-16", "Datos comerciales IA", "ventasDiarias BFF-only, ireHistorial extendido, grants lectura.", "MP-12, MP-08"),
    ("1.3.0", "2026-05-19", "Pedidos hardened", "RLS pedidos/usuarios, idempotency_key, admin sales atomic.", "MP-05, MP-07"),
    ("1.4.0", "2026-05-22", "Roles y auditoría", "tighten usuario_roles, harden auditoría service_role.", "MP-10, MP-06"),
    ("1.5.0", "2026-05-26", "Legal Perú", "libro_reclamaciones virtual, cookies, términos Ley 29571/29733.", "MP-15, MP-16"),
    ("1.6.0", "2026-05-31", "Supabase hardening", "linter remediation v1/v2, usuarios_seguro invoker, retención datos.", "MP-11, MP-10"),
    ("1.7.0", "2026-06-08", "IA campañas", "campanas_detectadas v2, modelo_estado, feedback humano campaña.", "MP-12, MP-14"),
    ("1.8.0", "2026-06-12", "Admin pedidos UX", "mapa entrega admin, buscador/paginación pedidos, datos cliente completos.", "MP-05, MP-04, MP-07"),
    ("1.9.0", "2026-06-16", "ISO 25000 + QC", "PKCS#7 no repudio, qc_* adecuación funcional, mantenibilidad 100%, portabilidad gates.", "MP-10, MP-11, MP-13"),
]

# Versión actual por macroproceso (última iteración web que lo tocó)
MP_VERSION: dict[str, tuple[str, str, str]] = {
    "MP-01": ("1.9.0", "2026-06-16", "Alcance SRS + trazabilidad ISO 25000 y documentación tesis integrada."),
    "MP-02": ("1.5.0", "2026-05-26", "Home, landings campañas y navegación pública estable."),
    "MP-03": ("1.7.0", "2026-06-08", "Catálogo, variantes, imágenes Cloudinary, campaña en productos."),
    "MP-04": ("1.8.0", "2026-06-12", "Carrito, checkout Stripe/COD, entrega geocodificada ORS."),
    "MP-05": ("1.9.0", "2026-06-16", "Pedidos cliente/admin, mapa, PKCS#7, búsqueda y paginación."),
    "MP-06": ("1.6.0", "2026-05-31", "Registro, perfil, favoritos, DNI pgcrypto, App Check."),
    "MP-07": ("1.8.0", "2026-06-12", "Dashboard admin, productos, stock, finanzas, fabricantes BFF."),
    "MP-08": ("1.3.0", "2026-05-19", "Ventas diarias, devoluciones, panel staff/trabajador."),
    "MP-09": ("1.2.0", "2026-05-16", "Import/export Excel, reglas comerciales, sanitización."),
    "MP-10": ("1.9.0", "2026-06-16", "RLS, BFF fail-closed, ZAP prod, rate limit, auditoría PII."),
    "MP-11": ("1.9.0", "2026-06-16", "64 migraciones, qc_* RLS, anti-regresión, modelo_estado."),
    "MP-12": ("1.7.0", "2026-06-08", "IRE, Random Forest, campañas IA, backtest gate, Render."),
    "MP-13": ("1.9.0", "2026-06-16", "CI/CD, Sonar, k6 mixed1000, restore drill, gates ISO."),
    "MP-14": ("1.7.0", "2026-06-08", "Landings campaña, detección IA, trazabilidad comercial."),
    "MP-15": ("1.5.0", "2026-05-26", "Libro reclamaciones virtual implementado."),
    "MP-16": ("1.5.0", "2026-05-26", "Política privacidad, cookies, cumplimiento Ley 29733."),
}

# Overrides finos por bolt (iteraciones posteriores al MP base)
BOLT_VERSION_OVERRIDE: dict[str, tuple[str, str, str]] = {
    "BOLT-WEB-012": ("1.8.0", "2026-06-12", "v1.8: listado admin pedidos con búsqueda, paginación y mapa entrega."),
    "BOLT-WEB-037": ("1.8.0", "2026-06-12", "v1.8: checkout entrega + rutas Google Maps / ORS."),
    "BOLT-WEB-047": ("1.8.0", "2026-06-12", "v1.8: detalle pedido admin con datos cliente y enlace mapa."),
    "BOLT-WEB-021": ("1.9.0", "2026-06-16", "v1.9: ZAP production v2 + guard tests seguridad."),
    "BOLT-WEB-025": ("1.9.0", "2026-06-16", "v1.9: CI con gates ISO 25000, E2E multi-browser, typecheck."),
    "BOLT-WEB-054": ("1.9.0", "2026-06-16", "v1.9: SonarQube + mantenibilidad 100% verificable."),
    "BOLT-WEB-018": ("1.9.0", "2026-06-16", "v1.9: auditoría admin trail + PII enforce at insert."),
    "BOLT-WEB-046": ("1.9.0", "2026-06-16", "v1.9: auditoría PII enmascarada en logs."),
    "BOLT-WEB-022": ("1.9.0", "2026-06-16", "v1.9: RLS matrix + migración qc_* + PKCS#7 pedidos."),
    "BOLT-WEB-044": ("1.9.0", "2026-06-16", "v1.9: anti-regresión migraciones Supabase."),
    "BOLT-WEB-045": ("1.9.0", "2026-06-16", "v1.9: service_role policies + linter remediation v2."),
    "BOLT-WEB-026": ("1.9.0", "2026-06-16", "v1.9: k6 mixed1000 BFF live + autocannon stress."),
    "BOLT-WEB-027": ("1.9.0", "2026-06-16", "v1.9: restore drill Supabase real + fiabilidad 84% prod."),
    "BOLT-WEB-052": ("1.9.0", "2026-06-16", "v1.9: continuidad ops documentada post drill."),
    "BOLT-WEB-031": ("1.9.0", "2026-06-16", "v1.9: portabilidad adaptabilidad/intercambiabilidad gates."),
    "BOLT-WEB-024": ("1.7.0", "2026-06-08", "v1.7: panel IRE + campañas + sparkline historial."),
    "BOLT-WEB-053": ("1.7.0", "2026-06-08", "v1.7: ai-backtest-gate + data_sufficient + model card."),
    "BOLT-WEB-028": ("1.7.0", "2026-06-08", "v1.7: campañas detectadas + feedback adaptativo."),
    "BOLT-WEB-055": ("1.7.0", "2026-06-08", "v1.7: landings públicas campaña comercial."),
    "BOLT-IA-001": ("1.7.0", "2026-06-08", "FastAPI combined + campaign endpoints Render."),
    "BOLT-IA-002": ("1.7.0", "2026-06-08", "IRE v1.2 fórmula 40/35/25 + proyección horizonte."),
    "BOLT-IA-003": ("1.6.0", "2026-05-31", "Random Forest demanda + fallback conservador."),
    "BOLT-IA-004": ("1.7.0", "2026-06-08", "modelo_estado + ireHistorial persistencia Supabase."),
}

SHEET_COLUMN_WIDTHS: dict[str, list[float]] = {
    "PROCESOS": [11, 14, 42, 22, 14, 12, 12, 52],
    "ACTIVIDADES_BPMN": [14, 11, 32, 18, 12, 58, 24],
    "FLUJO_BPMN": [12, 11, 22, 38, 14, 14, 28],
    "REQUERIMIENTOS": [16, 10, 62, 10, 10, 12, 16],
    "CONTROL_VERSIONES": [12, 14, 14, 10, 28, 12, 14, 12, 22, 28, 14, 10, 24, 8, 8, 48],
    "AI_DLC": [14, 48, 28, 32, 38, 38, 22, 10, 12],
    "DEPENDENCIAS": [14, 14, 16, 10, 52],
    "MERGES": [10, 16, 22, 12, 12, 22, 12, 48],
    "HISTORIAL": [12, 14, 10, 22, 58, 28, 22],
    "ARTEFACTOS": [10, 12, 14, 32, 42, 8, 48],
    "PMV": [10, 32, 14, 16, 12, 48],
    "METADATA": [28, 58],
    "RESPALDO_Q1": [22, 8, 36, 28, 38, 22, 42],
    "CONTROLES_ISO42001": [14, 16, 38, 38, 8, 38, 10, 42],
    "RIESGOS_IA": [10, 12, 42, 12, 10, 38, 12, 32, 8, 38],
    "MODEL_CARD": [22, 48, 32],
    "VERSIONES_WEB": [10, 12, 18, 52, 18],
    "DASHBOARD": [32, 18, 12],
}


def bolt_version_info(bolt_id: str, process: str) -> tuple[str, str, str]:
    if bolt_id in BOLT_VERSION_OVERRIDE:
        return BOLT_VERSION_OVERRIDE[bolt_id]
    mp = MP_VERSION.get(process, ("1.0.0", "2026-05-01", "Implementación inicial."))
    return mp


def build_macro_processes() -> list[tuple]:
    base = [
        ("MP-01", "Estratégico", "Dirección comercial y alcance del sistema", "feature/alcance-sistema", "Gestión"),
        ("MP-02", "Misional", "Experiencia pública y captación digital", "feature/experiencia-publica", "Operaciones"),
        ("MP-03", "Misional", "Gestión del catálogo comercial", "feature/catalogo", "Operaciones"),
        ("MP-04", "Misional", "Gestión de carrito y compra web", "feature/carrito-checkout", "Operaciones"),
        ("MP-05", "Misional", "Gestión de pedidos y posventa", "feature/pedidos", "Operaciones"),
        ("MP-06", "Misional", "Gestión de identidad, clientes y privacidad", "feature/clientes", "Operaciones"),
        ("MP-07", "Misional", "Administración comercial", "feature/admin-comercial", "Gestión"),
        ("MP-08", "Misional", "Gestión de ventas físicas y operación de tienda", "feature/ventas-fisicas", "Operaciones"),
        ("MP-09", "Apoyo", "Gestión de datos Excel e importación/exportación", "feature/admin-datos", "Soporte"),
        ("MP-10", "Apoyo", "Seguridad, auditoría y cumplimiento", "feature/seguridad", "Tecnología"),
        ("MP-11", "Apoyo", "Base de datos y control de acceso Supabase", "feature/supabase-rls", "Tecnología"),
        ("MP-12", "Misional", "Analítica, IA y riesgo empresarial (IRE)", "feature/ia-ire", "Analítica"),
        ("MP-13", "Apoyo", "DevOps, pruebas y continuidad", "feature/devops", "Tecnología"),
        ("MP-14", "Misional", "Campañas, marketing y trazabilidad comercial", "feature/campanas", "Operaciones"),
        ("MP-15", "Apoyo", "Atención al consumidor y libro de reclamaciones", "feature/libro-reclamaciones", "Soporte"),
        ("MP-16", "Apoyo", "Consentimiento, políticas y cumplimiento legal", "feature/legal", "Soporte"),
    ]
    rows = []
    for mp_id, tipo, nombre, rama, resp in base:
        ver, fecha, nota = MP_VERSION[mp_id]
        obs = f"Web {ver} ({fecha}). {nota}"
        if mp_id == "MP-12":
            obs += f" Gobierno IA plantilla {CONTENT_VERSION}."
        rows.append((mp_id, tipo, nombre, rama, resp, "PRODUCCION", ver, obs))
    return rows


MACRO_PROCESSES = build_macro_processes()

MACRO_BY_PREFIX = {
    "MP-01": "MP-01", "MP-02": "MP-02", "MP-03": "MP-03", "MP-04": "MP-04",
    "MP-05": "MP-05", "MP-06": "MP-06", "MP-07": "MP-07", "MP-08": "MP-08",
    "MP-09": "MP-09", "MP-10": "MP-10", "MP-11": "MP-11", "MP-12": "MP-12",
    "MP-13": "MP-13", "MP-14": "MP-14", "MP-15": "MP-15", "MP-16": "MP-16",
    "MP-17": "MP-13", "MP-18": "MP-03",
}

PROCESS_BY_BOLT = {
    "BOLT-WEB-001": "MP-01", "BOLT-WEB-002": "MP-02", "BOLT-WEB-003": "MP-03",
    "BOLT-WEB-004": "MP-03", "BOLT-WEB-005": "MP-04", "BOLT-WEB-006": "MP-04",
    "BOLT-WEB-007": "MP-05", "BOLT-WEB-008": "MP-06", "BOLT-WEB-009": "MP-06",
    "BOLT-WEB-010": "MP-07", "BOLT-WEB-011": "MP-07", "BOLT-WEB-012": "MP-05",
    "BOLT-WEB-013": "MP-07", "BOLT-WEB-014": "MP-08", "BOLT-WEB-015": "MP-08",
    "BOLT-WEB-016": "MP-09", "BOLT-WEB-017": "MP-09", "BOLT-WEB-018": "MP-10",
    "BOLT-WEB-019": "MP-04", "BOLT-WEB-020": "MP-10", "BOLT-WEB-021": "MP-10",
    "BOLT-WEB-022": "MP-11", "BOLT-WEB-023": "MP-11", "BOLT-WEB-024": "MP-12",
    "BOLT-WEB-025": "MP-13", "BOLT-WEB-026": "MP-13", "BOLT-WEB-027": "MP-13",
    "BOLT-WEB-028": "MP-14", "BOLT-WEB-029": "MP-15", "BOLT-WEB-030": "MP-16",
    "BOLT-WEB-031": "MP-13", "BOLT-WEB-032": "MP-06", "BOLT-WEB-033": "MP-06",
    "BOLT-WEB-034": "MP-05", "BOLT-WEB-035": "MP-07", "BOLT-WEB-036": "MP-03",
    "BOLT-WEB-037": "MP-04", "BOLT-WEB-038": "MP-09", "BOLT-WEB-039": "MP-09",
    "BOLT-WEB-040": "MP-09", "BOLT-WEB-041": "MP-07", "BOLT-WEB-042": "MP-03",
    "BOLT-WEB-043": "MP-07", "BOLT-WEB-044": "MP-11", "BOLT-WEB-045": "MP-11",
    "BOLT-WEB-046": "MP-11", "BOLT-WEB-047": "MP-05", "BOLT-WEB-048": "MP-10",
    "BOLT-WEB-049": "MP-08", "BOLT-WEB-050": "MP-06", "BOLT-WEB-051": "MP-10",
    "BOLT-WEB-052": "MP-13", "BOLT-WEB-053": "MP-12", "BOLT-WEB-054": "MP-13",
    "BOLT-WEB-055": "MP-14",
}

AI_SERVICE_BOLTS = [
    {
        "id": "BOLT-IA-001",
        "macro": "MP-12",
        "desc": "Exponer API FastAPI de predicción combinada, demanda, IRE y alertas de stock.",
        "domain": "ai-service / API",
        "location": "ai-service/main.py",
        "parent": "BOLT-WEB-024",
    },
    {
        "id": "BOLT-IA-002",
        "macro": "MP-12",
        "desc": "Calcular IRE 0-100 con dimensiones stock/ingresos/demanda y niveles de riesgo.",
        "domain": "ai-service / IRE",
        "location": "ai-service/models/risk.py",
        "parent": "BOLT-WEB-024",
    },
    {
        "id": "BOLT-IA-003",
        "macro": "MP-12",
        "desc": "Entrenar y ejecutar RandomForestRegressor para predicción de demanda por producto.",
        "domain": "ai-service / ML",
        "location": "ai-service/models/demand.py",
        "parent": "BOLT-WEB-024",
    },
    {
        "id": "BOLT-IA-004",
        "macro": "MP-12",
        "desc": "Persistir historial IRE, estado del modelo y lectura de ventas desde Supabase.",
        "domain": "ai-service / datos",
        "location": "ai-service/services/supabase_client.py",
        "parent": "BOLT-WEB-053",
    },
]

REQUIREMENTS = [
    ("RF-001", "RF", "El sistema debe mostrar catálogo público de calzado con filtros y detalle de producto.", "Alta", "PMV-01", "PRODUCCION", "BOLT-WEB-003"),
    ("RF-002", "RF", "El sistema debe gestionar carrito por sesión y checkout con validación de stock.", "Alta", "PMV-02", "PRODUCCION", "BOLT-WEB-005"),
    ("RF-003", "RF", "El sistema debe procesar pagos Stripe y contraentrega con confirmación de pedido.", "Alta", "PMV-02", "PRODUCCION", "BOLT-WEB-019"),
    ("RF-004", "RF", "El sistema debe registrar ventas físicas diarias, devoluciones y documentos de venta.", "Alta", "PMV-03", "PRODUCCION", "BOLT-WEB-014"),
    ("RF-005", "RF", "El sistema debe administrar productos, stock, finanzas y fabricantes desde panel admin.", "Alta", "PMV-04", "PRODUCCION", "BOLT-WEB-010"),
    ("RF-006", "RF", "El sistema debe calcular y mostrar el IRE con variables, fórmula y nivel de riesgo.", "Alta", "PMV-05", "PRODUCCION", "BOLT-WEB-024"),
    ("RF-007", "RF", "El sistema debe predecir demanda por producto y alertar quiebres de stock.", "Alta", "PMV-05", "PRODUCCION", "BOLT-IA-001"),
    ("RF-008", "RF", "El sistema debe importar y exportar datos comerciales mediante plantillas Excel.", "Media", "PMV-04", "PRODUCCION", "BOLT-WEB-016"),
    ("RNF-001", "RNF", "El sistema debe aplicar RLS en Supabase y operaciones sensibles solo vía BFF.", "Alta", "PMV-06", "PRODUCCION", "BOLT-WEB-022"),
    ("RNF-002", "RNF", "El servicio IA debe autenticarse con Bearer token y no exponer secretos al cliente.", "Alta", "PMV-05", "PRODUCCION", "BOLT-IA-001"),
    ("RNF-003", "RNF", "El pipeline CI/CD debe ejecutar lint, typecheck, tests y gates antes de despliegue.", "Alta", "PMV-06", "PRODUCCION", "BOLT-WEB-025"),
    ("RNF-004", "RNF", "El módulo IA debe registrar backtesting, suficiencia de datos y calidad del modelo.", "Alta", "PMV-05", "PRODUCCION", "BOLT-WEB-053"),
    ("RNF-005", "RNF", "La aplicación debe cumplir controles de seguridad OWASP ZAP sin alertas altas.", "Alta", "PMV-06", "PRODUCCION", "BOLT-WEB-021"),
    ("RNF-006", "RNF", "Los datos personales deben minimizarse, enmascararse y protegerse según Ley 29733.", "Alta", "PMV-06", "PRODUCCION", "BOLT-WEB-023"),
    ("RNF-007", "RNF", "El IRE debe ser explicable: fórmula 40/35/25, contribuciones y niveles documentados.", "Alta", "PMV-05", "PRODUCCION", "BOLT-IA-002"),
    ("RNF-008", "RNF", "Ante data_sufficient=false el sistema debe bloquear proyecciones ML y mostrar aviso operativo.", "Alta", "PMV-05", "PRODUCCION", "BOLT-WEB-053"),
    ("RNF-009", "RNF", "Debe registrarse historial IRE y modelo_estado con versionado y trazabilidad auditable.", "Alta", "PMV-05", "PRODUCCION", "BOLT-IA-004"),
    ("RNF-010", "RNF", "Decisiones admin sobre campañas IA requieren feedback humano (confirmar/rechazar detección).", "Media", "PMV-05", "PRODUCCION", "BOLT-WEB-024"),
]

PMV_ROWS = [
    ("PMV-01", "Catálogo público navegable", "BOLT-WEB-003", "Productos", "PRODUCCION", f"Web {MP_VERSION['MP-03'][0]} — catálogo, filtros, detalle, campaña en producto."),
    ("PMV-02", "Compra web end-to-end", "BOLT-WEB-006", "Carrito/Pedidos", "PRODUCCION", f"Web {MP_VERSION['MP-04'][0]} — carrito → checkout → Stripe/COD → entrega geocodificada."),
    ("PMV-03", "Ventas físicas y devoluciones", "BOLT-WEB-014", "Ventas", "PRODUCCION", f"Web {MP_VERSION['MP-08'][0]} — ventas diarias, devoluciones, panel staff."),
    ("PMV-04", "Panel administrativo comercial", "BOLT-WEB-010", "Admin", "PRODUCCION", f"Web {MP_VERSION['MP-07'][0]} — productos, stock, finanzas, Excel, dashboard."),
    ("PMV-05", "IA comercial y gobierno del IRE", "BOLT-WEB-024", "IA", "PRODUCCION", f"Web {MP_VERSION['MP-12'][0]} + gobierno {CONTENT_VERSION} — IRE, RF, campañas, backtest."),
    ("PMV-06", "Seguridad, calidad y DevOps", "BOLT-WEB-025", "Operaciones", "PRODUCCION", f"Web {WEB_APP_VERSION} — CI/CD, RLS, ZAP, k6, gates ISO 25000, Sonar."),
]

BPMN_IA_FLOW = [
    ("N-IA-01", "MP-12", "EVENTO", "Inicio solicitud de predicción", "", "N-IA-02", ""),
    ("N-IA-02", "MP-12", "ACTIVIDAD", "Recolectar ventasDiarias y pedidos pagados", "N-IA-01", "N-IA-03", ""),
    ("N-IA-03", "MP-12", "GATEWAY", "¿Datos suficientes para modelo?", "N-IA-02", "N-IA-04", "Sí"),
    ("N-IA-04", "MP-12", "ACTIVIDAD", "Entrenar/ejecutar Random Forest de demanda", "N-IA-03", "N-IA-05", ""),
    ("N-IA-05", "MP-12", "ACTIVIDAD", "Calcular forecast de ingresos", "N-IA-04", "N-IA-06", ""),
    ("N-IA-06", "MP-12", "ACTIVIDAD", "Calcular IRE y alertas de stock", "N-IA-05", "N-IA-07", ""),
    ("N-IA-07", "MP-12", "ACTIVIDAD", "Persistir ireHistorial y modelo_estado", "N-IA-06", "N-IA-08", ""),
    ("N-IA-08", "MP-12", "ACTIVIDAD", "Mostrar dashboard admin de predicciones", "N-IA-07", "N-IA-09", ""),
    ("N-IA-09", "MP-12", "EVENTO", "Fin — decisión comercial asistida", "N-IA-08", "", ""),
    ("N-IA-03B", "MP-12", "ACTIVIDAD", "Emitir aviso data_sufficient=false (cold start)", "N-IA-03", "N-IA-09", "No"),
]

MERGES = [
    ("MRG-001", "release → main", "release/1.0-bootstrap", "main", "2026-05-01", OWNER, "INTEGRADO", "v1.0.0 Catálogo, auth, pedidos base."),
    ("MRG-002", "release → main", "release/1.1-stock", "main", "2026-05-14", OWNER, "INTEGRADO", "v1.1.0 Movimientos stock y ventas admin."),
    ("MRG-003", "release → main", "release/1.2-ia-data", "main", "2026-05-16", OWNER, "INTEGRADO", "v1.2.0 ventasDiarias + ireHistorial."),
    ("MRG-004", "release → main", "release/1.3-pedidos", "main", "2026-05-19", OWNER, "INTEGRADO", "v1.3.0 RLS pedidos e idempotency."),
    ("MRG-005", "release → main", "release/1.5-legal", "main", "2026-05-26", OWNER, "INTEGRADO", "v1.5.0 Libro reclamaciones y legal."),
    ("MRG-006", "release → main", "release/1.7-ia-campanas", "main", "2026-06-08", OWNER, "INTEGRADO", "v1.7.0 IA campañas + modelo_estado."),
    ("MRG-007", "release → main", "release/1.8-admin-pedidos", "main", "2026-06-12", OWNER, "INTEGRADO", "v1.8.0 Mapas entrega y UX pedidos admin."),
    ("MRG-008", "release → main", "release/1.9-iso25000", "main", "2026-06-16", OWNER, "INTEGRADO", "v1.9.0 PKCS#7, qc_*, gates ISO 25000."),
    ("MRG-009", "hotfix → main", "fix/bff-manufacturers", "main", "2026-05-28", OWNER, "INTEGRADO", "Fabricantes vía BFF sin Supabase directo."),
]

HISTORY = [
    (TODAY, "PLANTILLA", CONTENT_VERSION, "Gobierno IA V3", f"Sync web {WEB_APP_VERSION}: versiones por bolt, VERSIONES_WEB, formato celdas.", AGENT, OWNER),
    ("2026-06-16", "BOLT-WEB-022", "1.9.0", "ISO + QC Supabase", "Migraciones qc_* + PKCS#7 + RLS service_role.", AGENT, OWNER),
    ("2026-06-16", "BOLT-WEB-025", "1.9.0", "CI ISO 25000", "Gates mantenibilidad 100%, portabilidad, E2E multi-browser.", AGENT, OWNER),
    ("2026-06-16", "BOLT-WEB-021", "1.9.0", "Seguridad DAST", "ZAP production v2 sin alertas altas.", AGENT, OWNER),
    ("2026-06-12", "BOLT-WEB-012", "1.8.0", "Admin pedidos", "Buscador, paginación, mapa Google Maps en detalle.", AGENT, OWNER),
    ("2026-06-12", "BOLT-WEB-037", "1.8.0", "Checkout entrega", "Geocodificación y rutas ORS/Nominatim.", AGENT, OWNER),
    ("2026-06-08", "BOLT-WEB-024", "1.7.0", "Panel IRE v2", "Historial sparkline, campañas IA, tabs dashboard.", AGENT, OWNER),
    ("2026-06-08", "BOLT-WEB-053", "1.7.0", "Gobierno modelo", "ai-backtest-gate, data_sufficient, model card.", AGENT, OWNER),
    ("2026-06-08", "BOLT-IA-001", "1.7.0", "API campañas", "Endpoints campaign-detection y feedback.", AGENT, OWNER),
    ("2026-05-31", "BOLT-WEB-045", "1.6.0", "Linter Supabase", "Remediation security invoker + policies.", AGENT, OWNER),
    ("2026-05-26", "BOLT-WEB-029", "1.5.0", "Legal Perú", "Libro reclamaciones virtual Ley 29571.", AGENT, OWNER),
    ("2026-05-19", "BOLT-WEB-014", "1.3.0", "Ventas admin", "Atomic hardening ventas y devoluciones.", AGENT, OWNER),
    ("2026-05-16", "BOLT-IA-004", "1.7.0", "Persistencia IA", "ireHistorial extend + modelo_estado table.", AGENT, OWNER),
    ("2026-05-14", "BOLT-WEB-011", "1.1.0", "Stock RPC", "movimientos_stock + registrar ingreso.", AGENT, OWNER),
    ("2026-05-01", "BOLT-WEB-003", "1.0.0", "Bootstrap catálogo", "Catálogo público React + Supabase.", AGENT, OWNER),
    ("2026-05-28", "PLANTILLA", "1.0", "Gobierno IA V3", "Generación inicial plantilla desde bolts.", AGENT, OWNER),
    ("2026-06-19", "PLANTILLA", "1.1", "Ampliación Q1", "RESPALDO_Q1, CONTROLES_ISO42001, RIESGOS_IA, MODEL_CARD.", AGENT, OWNER),
]

# ISO/IEC 42001 — controles operacionalizados (evidencia repo + artículo Q1)
CONTROLES_ISO42001 = [
    ("CTL-42001-01", "4 Contexto", "Alcance AIMS: e-commerce + IRE comercial-operativo PYME", "01-marco-y-tesis.md; formato-09", "01", "10.1016/j.jbusres.2019.09.022", "VERDE", "Verhoef 2019 — transformación digital"),
    ("CTL-42001-02", "5 Liderazgo", "Roles admin/dirección para decisiones asistidas por IA", "AdminPredictions.tsx; RF-IA-03", "11", "10.1177/0008125619862257", "VERDE", "Shrestha 2019 — decisiones organizacionales IA"),
    ("CTL-42001-03", "6 Planificación", "Riesgos IA documentados (drift, cold start, sesgo pesos)", "07-modulo-ia §9; hoja RIESGOS_IA", "41", "10.1007/s10664-021-09993-1", "VERDE", "Haakman 2021 — ciclo vida IA"),
    ("CTL-42001-04", "7 Soporte — Datos", "Calidad y suficiencia data_sufficient + ventasDiarias", "ai-service/models/demand/predict.py; predictionDataQuality.test.ts", "40", "10.1111/poms.12838", "VERDE", "Choi 2018 — analytics operaciones"),
    ("CTL-42001-05", "7 Soporte — Datos", "Big data KPIs alimentan IRE", "RF-ADM-01; ireHistorial", "12", "10.1016/j.jbusres.2016.08.009", "VERDE", "Amba 2017 — big data desempeño"),
    ("CTL-42001-06", "8 Operación — Ciclo vida", "CRISP-ML(Q): entrenar → desplegar → monitorear", "ai-service/; modelo_estado; evaluate.py", "29", "10.1016/j.infsof.2020.106368", "VERDE", "Lwakatare 2020 — ML industrial"),
    ("CTL-42001-07", "8 Operación — Despliegue", "Desafíos ML producción mitigados (FastAPI Render Docker)", "ai-service/Dockerfile; docs/ops", "32", "10.1145/3533378", "AMBAR", "Paleyes 2022 — despliegue ML; series PYME cortas"),
    ("CTL-42001-08", "8 Operación — SE for AI", "Ingeniería sistemas basados en IA verificable", "verify-idoneidad; Martínez-Fernández trazabilidad", "33", "10.1145/3487043", "VERDE", "TOSEM 2022 — SE for AI"),
    ("CTL-42001-09", "8 Operación — Validación", "Backtesting MAPE/RMSE gate automatizado", "scripts/ai-backtest-gate.mjs; evaluate.py", "13", "10.1371/journal.pone.0194889", "VERDE", "Makridakis 2018 — forecasting"),
    ("CTL-42001-10", "8 Operación — Modelo", "Random Forest demanda con fallback conservador", "ai-service/models/demand.py", "35", "10.1023/A:1010933404324", "VERDE", "Breiman 2001 — random forests"),
    ("CTL-42001-11", "8 Operación — Explicabilidad", "IRE reglas explícitas 40/35/25 + contribuciones UI", "ai-service/models/risk.py; AdminPredictions", "09", "10.1016/j.ijinfomgt.2019.01.021", "VERDE", "Duan 2019 — IA decisiones Big Data"),
    ("CTL-42001-12", "8 Operación — Prescripción", "Alertas stock y recomendaciones ABC", "IN-27; IN-34; RF-IA-03", "39", "10.1016/j.ijinfomgt.2019.04.003", "VERDE", "Lepenioti 2020 — prescriptive analytics"),
    ("CTL-42001-13", "8 Operación — Supervisión humana", "Feedback campaña confirmar/rechazar + panel admin", "ai-service/models/campaign.py; POST /api/campaign/feedback", "11", "10.1177/0008125619862257", "VERDE", "Human-in-the-loop campañas"),
    ("CTL-42001-14", "8 Operación — Monitoreo", "drift_score + historial IRE longitudinal", "ireHistorial; test_risk.py proyección", "14", "10.1016/j.ijforecast.2019.06.004", "AMBAR", "Fildes 2019 — retail forecasting"),
    ("CTL-42001-15", "8 Operación — Seguridad IA", "Auth Bearer Firebase; BFF fail-closed; RLS", "firebase_verifier.py; verify-seguridad", "44", "10.1016/j.cose.2021.102436", "VERDE", "Buck 2021 — zero trust"),
    ("CTL-42001-16", "8 Operación — Ciber e-commerce", "DAST ZAP + hardening comercio digital", "zap-production-report-v2.json", "47", "10.1016/j.cose.2021.102248", "VERDE", "Lallie 2021 — ciberataques e-commerce"),
    ("CTL-42001-17", "9 Evaluación desempeño", "Métricas modelo + gates CI antes release", "ci.yml; ai-backtest-gate.mjs", "33", "10.1145/3487043", "VERDE", "Evaluación continua"),
    ("CTL-42001-18", "10 Mejora", "Reentrenamiento bajo demanda + versionado modelo_estado", "save_modelo_estado; migrations", "41", "10.1007/s10664-021-09993-1", "VERDE", "Mejora ciclo IA"),
    ("CTL-42001-19", "Contexto PYME", "E-commerce + IA integrado vs SaaS genérico", "Matriz propuesta valor; CU-T12", "43", "10.1371/journal.pone.0305639", "VERDE", "Dai 2024 — IA big data PYME"),
    ("CTL-42001-20", "Arquitectura", "Microservicios BFF + Supabase + ai-service", "arquitectura-sistema.md; Jamshidi trazabilidad", "28", "10.1109/ms.2018.2141039", "VERDE", "Jamshidi 2018 — microservicios"),
]

RIESGOS_IA = [
    ("R-IA-01", "Datos", "Historial corto PYME limita generalización ML", "Alta", "Media", "data_sufficient + fallback promedio móvil", "Mitigado", "07-modulo-ia §9", "32", "10.1145/3533378"),
    ("R-IA-02", "Modelo", "Drift por cambios de catálogo/temporada", "Media", "Media", "drift_score + recálculo periódico", "Monitoreado", "models/demand.py", "14", "10.1016/j.ijforecast.2019.06.004"),
    ("R-IA-03", "Validación", "Sin etiquetas de crisis → no AUC del IRE", "Alta", "Baja", "Validación consistencia + sensibilidad + Likert O₂", "Aceptado", "07-modulo-ia §7.2", "09", "10.1016/j.ijinfomgt.2019.01.021"),
    ("R-IA-04", "Gobernanza", "Pesos IRE 40/35/25 percibidos como arbitrarios", "Media", "Media", "Sensibilidad documentada + aprobación asesor", "Mitigado", "07-modulo-ia §4", "11", "10.1177/0008125619862257"),
    ("R-IA-05", "Operación", "Cold start: predicciones sin datos suficientes", "Alta", "Alta", "Banner data_sufficient=false; E2E TC-PRED-003", "Controlado", "admin-predictions.spec.ts", "41", "10.1007/s10664-021-09993-1"),
    ("R-IA-06", "Seguridad", "Exposición API IA sin autenticación", "Crítica", "Baja", "Firebase Bearer + tests contrato API", "Mitigado", "test_firebase_verifier.py", "44", "10.1016/j.cose.2021.102436"),
    ("R-IA-07", "Sesgo", "Sesgo hacia productos alta rotación en alertas", "Media", "Media", "Análisis ABC + revisión admin", "Monitoreado", "IN-34", "40", "10.1111/poms.12838"),
    ("R-IA-08", "Legal", "Decisiones automatizadas sin política explícita IA", "Media", "Baja", "Términos + política privacidad + supervisión humana", "Mitigado", "RF-LEG; Ley 29733", "44", "10.1016/j.cose.2021.102436"),
]

MODEL_CARD = [
    ("Nombre", "Calzatura Vilchez — IRE + Demand Forecast", "07-modulo-ia"),
    ("Versión modelo", f"{WEB_APP_VERSION} / gobierno {CONTENT_VERSION}", "modelo_estado.version"),
    ("Tipo", "RandomForestRegressor (demanda) + índice compuesto IRE (reglas)", "ai-service/models/"),
    ("Propósito", "Predicción demanda/ingresos y riesgo comercial-operativo 0-100", "Título tesis"),
    ("Datos entrenamiento", "ventasDiarias, pedidos pagados, productos (Supabase)", "§5 07-modulo-ia"),
    ("Métricas", "MAPE, RMSE backtest; gates ai-backtest-gate.mjs", "evaluate.py"),
    ("Limitaciones", "No predice quiebra financiera Altman; alcance PYME Huancayo", "§1 matiz IRE"),
    ("Supervisión humana", "Admin confirma campañas; dirección interpreta IRE", "campaign feedback API"),
    ("Actualización", "Reentrenamiento bajo demanda; historial ireHistorial", "save_modelo_estado"),
    ("Contacto", "Serpa Sedano Yeferson Wilson / Calzatura Vilchez", "CU-T01"),
]

RESPALDO_Q1 = [
    ("Gobierno / ciclo vida IA", "41", "Haakman et al. (2021)", "Empirical Software Engineering", "10.1007/s10664-021-09993-1", "Haakman 2021", "Marco AI-DLC y controles ISO 42001"),
    ("Despliegue ML producción", "32", "Paleyes et al. (2022)", "ACM Computing Surveys", "10.1145/3533378", "Paleyes 2022", "Retos MLOps y despliegue"),
    ("ML industrial PYME", "29", "Lwakatare et al. (2020)", "Information and Software Technology", "10.1016/j.infsof.2020.106368", "Lwakatare 2020", "Pipeline CI/CD IA"),
    ("Ingeniería software IA", "33", "Martínez-Fernández et al. (2022)", "ACM TOSEM", "10.1145/3487043", "Martínez-Fernández 2022", "Calidad sistemas IA"),
    ("IA decisiones Big Data", "09", "Duan et al. (2019)", "International Journal of Information Management", "10.1016/j.ijinfomgt.2019.01.021", "Duan 2019", "Explicabilidad y decisiones"),
    ("Forecasting / validación", "13", "Makridakis et al. (2018)", "PLOS ONE", "10.1371/journal.pone.0194889", "Makridakis 2018", "Backtesting y benchmarks"),
    ("Random Forest", "35", "Breiman (2001)", "Machine Learning", "10.1023/A:1010933404324", "Breiman 2001", "Algoritmo demanda"),
    ("Big data operaciones", "40", "Choi et al. (2018)", "Production and Operations Management", "10.1111/poms.12838", "Choi 2018", "Calidad datos operativos"),
    ("Big data desempeño", "12", "Amba et al. (2017)", "Journal of Business Research", "10.1016/j.jbusres.2016.08.009", "Amba 2017", "Analytics PYME"),
    ("Retail forecasting", "14", "Fildes et al. (2019)", "International Journal of Forecasting", "10.1016/j.ijforecast.2019.06.004", "Fildes 2019", "Monitoreo demanda retail"),
    ("Prescriptive analytics", "39", "Lepenioti et al. (2020)", "International Journal of Information Management", "10.1016/j.ijinfomgt.2019.04.003", "Lepenioti 2020", "Alertas y recomendaciones"),
    ("Adopción organizacional IA", "11", "Shrestha et al. (2019)", "California Management Review", "10.1177/0008125619862257", "Shrestha 2019", "Supervisión humana"),
    ("E-commerce + IA PYME", "43", "Dai et al. (2024)", "PLOS ONE", "10.1371/journal.pone.0305639", "Dai 2024", "Contexto negocio"),
    ("Seguridad zero-trust", "44", "Buck et al. (2021)", "Computers & Security", "10.1016/j.cose.2021.102436", "Buck 2021", "Auth API IA / RLS"),
    ("Ciber e-commerce", "47", "Lallie et al. (2021)", "Computers & Security", "10.1016/j.cose.2021.102248", "Lallie 2021", "Hardening DAST"),
    ("Microservicios", "28", "Jamshidi et al. (2018)", "IEEE Software", "10.1109/ms.2018.2141039", "Jamshidi 2018", "Arquitectura ai-service"),
    ("Transformación digital", "01", "Verhoef et al. (2019)", "Journal of Business Research", "10.1016/j.jbusres.2019.09.022", "Verhoef 2019", "Alcance AIMS"),
]

AI_DLC_DETAIL = {
    "BOLT-WEB-024": {
        "requerimiento": "RF-006/007; HU-IA-001; 07-modulo-ia v1.2",
        "diseno": "CRISP-ML(Q) + IRE 40/35/25 — Haakman 41; Duan 09",
        "desarrollo": "AdminPredictions.tsx + BFF proxy /api/predict/combined",
        "qa": "admin-predictions.spec.ts; admin-ire-dashboard.spec.ts; pytest test_risk",
        "produccion": "Firebase Hosting + Render ai-service — VERDE",
    },
    "BOLT-WEB-053": {
        "requerimiento": "RNF-004/008; ISO 42001 CTL-42001-09",
        "diseno": "Gate backtest MAPE — Makridakis 13; Choi 40 data quality",
        "desarrollo": "scripts/ai-backtest-gate.mjs; predictionDataQuality.test.ts",
        "qa": "node ai-backtest-gate.mjs --run; Vitest data_sufficient",
        "produccion": "CI ops + docs/ops/controles-devops-ia.md",
    },
    "BOLT-IA-001": {
        "requerimiento": "RF-007; contrato API OpenAPI",
        "diseno": "Microservicio FastAPI — Jamshidi 28; Paleyes 32",
        "desarrollo": "ai-service/main.py endpoints predict/combined",
        "qa": "test_api_contract.py; test_firebase_verifier.py",
        "produccion": "Docker Render; health check",
    },
    "BOLT-IA-002": {
        "requerimiento": "RF-006; RNF-007 explicabilidad",
        "diseno": "IRE reglas — Duan 09; NO Altman quiebra",
        "desarrollo": "ai-service/models/risk.py compute_ire",
        "qa": "test_risk.py invariantes pesos/umbrales/contribuciones",
        "produccion": "ireHistorial persistido",
    },
    "BOLT-IA-003": {
        "requerimiento": "RF-007 demanda por producto",
        "diseno": "Random Forest — Breiman 35; fallback conservador",
        "desarrollo": "ai-service/models/demand.py",
        "qa": "test_demand.py; evaluate.py backtest",
        "produccion": "modelo_estado + data_sufficient flag",
    },
    "BOLT-IA-004": {
        "requerimiento": "RNF-009 trazabilidad",
        "diseno": "ISO 25012 + 42001 datos — Amba 12; Lwakatare 29",
        "desarrollo": "supabase_client.py; ire_model.py",
        "qa": "test_supabase_client.py mocks",
        "produccion": "Migraciones ireHistorial/modeloEstado",
    },
}


def load_web_bolts() -> list[dict]:
    wb = openpyxl.load_workbook(BOLT_MATRIX, data_only=True)
    ws = wb["Bolts"]
    bolts = []
    for row in ws.iter_rows(min_row=5, max_col=78, values_only=True):
        if not row[1]:
            continue
        macro_raw = str(row[0] or "")
        macro_code = macro_raw.split()[0] if macro_raw else "MP-01"
        macro_code = MACRO_BY_PREFIX.get(macro_code, macro_code)
        bolts.append({
            "id": str(row[1]),
            "macro": macro_code,
            "desc": str(row[2] or ""),
            "domain": str(row[5] or ""),
            "fecha_matrix": str(row[4] or ""),
            "status": str(row[74] or "VERDE"),
            "observation": str(row[75] or ""),
            "location": str(row[77] or ""),
            "process": PROCESS_BY_BOLT.get(str(row[1]), macro_code),
        })
    return bolts


def clear_data_rows(ws, start_row: int = 2):
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def write_rows(ws, rows: list[tuple], start_row: int = 2):
    for offset, row in enumerate(rows):
        excel_row = start_row + offset
        ws.row_dimensions[excel_row].height = 42
        for col, value in enumerate(row, start=1):
            cell = ws.cell(excel_row, col, value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(name="Arial", size=9)


def apply_sheet_format(ws, sheet_name: str) -> None:
    widths = SHEET_COLUMN_WIDTHS.get(sheet_name)
    if not widths:
        return
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 36
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill("solid", fgColor="F5F9FC")
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    max_col = len(widths)
    max_row = ws.max_row or 1
    for row_idx in range(1, max_row + 1):
        if row_idx > 1 and row_idx % 2 == 0:
            row_fill = alt_fill
        else:
            row_fill = None
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.border = border
            if row_idx == 1:
                continue
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(name="Arial", size=9)
            if row_fill and col_idx <= max_col:
                cell.fill = row_fill
            if row_idx > 1:
                ws.row_dimensions[row_idx].height = max(ws.row_dimensions[row_idx].height or 0, 42)


def format_all_sheets(wb) -> None:
    for name in wb.sheetnames:
        if name == "DASHBOARD":
            ws = wb[name]
            ws.column_dimensions["A"].width = 34
            ws.column_dimensions["B"].width = 22
            ws.column_dimensions["C"].width = 14
            for row in range(1, 14):
                ws.row_dimensions[row].height = 22
                for col in ("A", "B", "C"):
                    c = ws[f"{col}{row}"]
                    if c.value:
                        c.alignment = Alignment(wrap_text=True, vertical="center")
                        c.font = Font(name="Arial", size=10)
            continue
        style_header_row(ws := wb[name], ws.max_column)
        apply_sheet_format(ws, name)


def ai_dlc_row(bolt_id: str, intent: str, stages: dict[str, str]) -> tuple:
    return (
        bolt_id,
        intent,
        stages.get("requerimiento", "Completado"),
        stages.get("diseno", "Completado"),
        stages.get("desarrollo", "Completado"),
        stages.get("qa", "Completado"),
        stages.get("produccion", "Completado"),
        stages.get("resultado", "VERDE"),
        TODAY,
    )


def build_dependencies(bolts: list[dict]) -> list[tuple]:
    deps = [
        ("BOLT-WEB-024", "BOLT-WEB-014", "Datos", "Alta", "IRE requiere ventasDiarias históricas."),
        ("BOLT-WEB-024", "BOLT-WEB-006", "Datos", "Alta", "Pedidos pagados alimentan demanda online."),
        ("BOLT-WEB-024", "BOLT-IA-001", "Servicio", "Alta", "UI admin consume API FastAPI."),
        ("BOLT-WEB-053", "BOLT-IA-001", "Gobierno", "Alta", "Controles de calidad sobre servicio IA."),
        ("BOLT-IA-002", "BOLT-IA-003", "Modelo", "Alta", "IRE usa salidas de predicción de demanda."),
        ("BOLT-IA-001", "BOLT-IA-004", "Datos", "Alta", "API depende de lectura Supabase."),
        ("BOLT-WEB-006", "BOLT-WEB-005", "Flujo", "Media", "Checkout parte del carrito activo."),
        ("BOLT-WEB-019", "BOLT-WEB-006", "Pago", "Alta", "Stripe/contraentrega tras checkout."),
        ("BOLT-WEB-025", "BOLT-WEB-024", "Calidad", "Media", "CI valida módulo IA antes de release."),
    ]
    return deps


def build_artifacts(bolts: list[dict]) -> list[tuple]:
    artifacts = [
        ("ART-001", "Documento", "BOLT-WEB-001", "SRS Calzatura Vilchez", "calzatura-vilchez/docs/01-srs/SRS-Calzatura-Vilchez.md", "1.0", "Requisitos funcionales y no funcionales."),
        ("ART-002", "Documento", "BOLT-WEB-024", "Módulo IA y IRE", "documentacion/07-modulo-ia-riesgo-empresarial.md", "1.2", "Definición operativa IRE v1.2."),
        ("ART-003", "Código", "BOLT-IA-001", "API FastAPI IA", "ai-service/main.py", "1.0", "Endpoints predict/combined."),
        ("ART-004", "Código", "BOLT-IA-002", "Motor IRE", "ai-service/models/risk.py", "1.0", "compute_ire y niveles."),
        ("ART-005", "Código", "BOLT-IA-003", "Modelo demanda RF", "ai-service/models/demand.py", "1.0", "RandomForestRegressor."),
        ("ART-006", "Código", "BOLT-WEB-024", "Panel predicciones", "calzatura-vilchez/src/domains/administradores/pages/AdminPredictions.tsx", "1.0", "UI IRE y métricas."),
        ("ART-007", "Código", "BOLT-WEB-003", "Catálogo web", "calzatura-vilchez/src/domains/productos/pages/ProductsPage.tsx", "1.0", "Catálogo público."),
        ("ART-008", "Código", "BOLT-WEB-014", "Ventas físicas", "calzatura-vilchez/src/domains/ventas/", "1.0", "Registro ventas y devoluciones."),
        ("ART-009", "Código", "BOLT-WEB-022", "BFF servidor", "calzatura-vilchez/bff/server.cjs", "1.0", "API segura y rate limit."),
        ("ART-010", "Evidencia", "BOLT-WEB-021", "Reporte ZAP", "zap-reports/zap-production-report-v2.json", "2.0", "DAST sin alertas altas."),
        ("ART-011", "Evidencia", "BOLT-WEB-026", "Stress test 2000", "artifacts/load-tests/autocannon-home-c2000-20260528-231905.json", "1.0", "Carga concurrente home."),
        ("ART-012", "Evidencia", "BOLT-WEB-053", "Backtest gate IA", "scripts/ai-backtest-gate.mjs", "1.0", "Control gobierno IA."),
        ("ART-013", "Matriz", "BOLT-WEB-001", "Estado del arte 43 tablas", "Estado_del_Arte_43_Tablas_CORREGIDO.docx", "1.0", "43 artículos con evidencia funcional."),
        ("ART-014", "Matriz", "BOLT-WEB-001", "Registro BOLT empresarial", "artifacts/matrices/Matriz_Registro_Web_Empresarial_AI_DLC_BOLT.xlsx", "1.0", "55 bolts web verificados."),
        ("ART-015", "Matriz", "BOLT-WEB-053", "Plantilla Gobierno IA V3", "artifacts/matrices/Plantilla_Gobierno_IA_V3_CALZATURA_VILCHEZ.xlsx", CONTENT_VERSION, "ISO 42001 + Q1 + model card."),
        ("ART-016", "Documento", "BOLT-WEB-053", "Guía gobierno IA", "documentacion/gobierno-ia-v3-calzatura-vilchez.md", CONTENT_VERSION, "Revisión y trazabilidad Q1."),
    ]
    for idx, b in enumerate(bolts[:8], start=15):
        if b["location"]:
            first_loc = b["location"].split(";")[0].strip()
            artifacts.append((
                f"ART-{idx:03d}",
                "Código",
                b["id"],
                b["domain"][:40],
                first_loc,
                "1.0",
                b["desc"][:80],
            ))
    return artifacts


def style_header_row(ws, ncol: int, fill: str = "1F4E79") -> None:
    header_fill = PatternFill("solid", fgColor=fill)
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for col in range(1, ncol + 1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def ensure_sheet(wb, name: str, headers: list[str]) -> None:
    if name in wb.sheetnames:
        ws = wb[name]
        clear_data_rows(ws)
    else:
        ws = wb.create_sheet(name)
        for col, header in enumerate(headers, start=1):
            ws.cell(1, col, header)
    style_header_row(ws, len(headers))
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


def write_metadata_sheet(wb) -> None:
    name = "METADATA"
    headers = ["Campo", "Valor"]
    ensure_sheet(wb, name, headers)
    ws = wb[name]
    rows = [
        ("Plantilla estructura", PLANTILLA_VERSION),
        ("Contenido Calzatura Vilchez", CONTENT_VERSION),
        ("Versión sistema web", WEB_APP_VERSION),
        ("Fecha generación", TODAY),
        ("Documento IA referencia", "07-modulo-ia-riesgo-empresarial.md v1.2"),
        ("Norma gobierno IA", "ISO/IEC 42001 (AIMS operacionalizado)"),
        ("Norma calidad datos", "ISO/IEC 25012 (referencia)"),
        ("Migraciones Supabase", "64 archivos en calzatura-vilchez/supabase/migrations/"),
        ("Corpus Q1", "43 artículos + complementos 44-48"),
        ("Regenerar", "python scripts/generar_gobierno_ia_v3.py"),
        ("Responsable", OWNER),
    ]
    write_rows(ws, rows)


def main():
    if not TEMPLATE.exists():
        raise FileNotFoundError(f"No se encontró plantilla: {TEMPLATE} ni {OUT_ARTIFACTS}")
    if not BOLT_MATRIX.exists():
        raise FileNotFoundError(f"No se encontró la matriz BOLT: {BOLT_MATRIX}")

    web_bolts = load_web_bolts()
    all_bolts = web_bolts + [
        {**b, "desc": b["desc"], "domain": b["domain"], "status": "VERDE",
         "observation": "Componente ai-service en producción Render.",
         "location": b["location"], "process": b["macro"]}
        for b in AI_SERVICE_BOLTS
    ]

    wb = openpyxl.load_workbook(TEMPLATE)

    # PROCESOS
    ws = wb["PROCESOS"]
    clear_data_rows(ws)
    write_rows(ws, build_macro_processes())

    # ACTIVIDADES_BPMN
    ws = wb["ACTIVIDADES_BPMN"]
    clear_data_rows(ws)
    act_rows = []
    for b in all_bolts:
        ver, fecha, nota = bolt_version_info(b["id"], b["process"])
        act_rows.append((
            b["id"],
            b["process"],
            f"Implementar {b['domain']} (web {ver})",
            "Desarrollador / Admin" if "IA" in b["domain"] or "Admin" in b["domain"] else "Desarrollador",
            "Tarea",
            f"{b['desc']} | Iteración {ver} ({fecha}): {nota}",
            b["domain"],
        ))
    write_rows(ws, act_rows)

    # FLUJO_BPMN
    ws = wb["FLUJO_BPMN"]
    clear_data_rows(ws)
    write_rows(ws, [(r[0], r[1], r[2], r[3], r[4], r[5], r[6]) for r in BPMN_IA_FLOW])

    # REQUERIMIENTOS
    ws = wb["REQUERIMIENTOS"]
    clear_data_rows(ws)
    write_rows(ws, REQUIREMENTS)

    # CONTROL_VERSIONES
    ws = wb["CONTROL_VERSIONES"]
    clear_data_rows(ws)
    cv_rows = []
    for idx, b in enumerate(all_bolts, start=1):
        estado = "PRODUCCION" if b["status"] == "VERDE" else "ACTIVO"
        ver, fecha, nota = bolt_version_info(b["id"], b["process"])
        cv_rows.append((
            f"CV-{idx:03d}",
            b["id"],
            AI_SERVICE_BOLTS[0]["parent"] if b["id"].startswith("BOLT-IA") else "",
            ver,
            f"Web {ver} — {nota}",
            estado,
            "INTEGRADO",
            fecha,
            OWNER,
            AGENT,
            f"main@{ver}",
            "",
            "Repositorio Calzatura Vilchez V3",
            "Alta",
            "No",
            b["observation"] if b.get("observation") else b["desc"],
        ))
    write_rows(ws, cv_rows)

    # AI_DLC
    ws = wb["AI_DLC"]
    clear_data_rows(ws)
    ai_bolts = [b for b in all_bolts if "IA" in b["id"] or "IA" in b["domain"] or b["id"] in {"BOLT-WEB-024", "BOLT-WEB-053"}]
    ai_rows = []
    for b in ai_bolts:
        detail = AI_DLC_DETAIL.get(b["id"], {})
        ai_rows.append(ai_dlc_row(
            b["id"],
            b["desc"][:120],
            {
                "requerimiento": detail.get("requerimiento", "HU-IA-001 / RF-006"),
                "diseno": detail.get("diseno", "CRISP-ML + 07-modulo-ia v1.2"),
                "desarrollo": detail.get("desarrollo", b.get("location", b["domain"]).split(";")[0]),
                "qa": detail.get("qa", "pytest + E2E + ai-backtest-gate"),
                "produccion": detail.get("produccion", "Render/Docker VERDE"),
                "resultado": "VERDE",
            },
        ))
    write_rows(ws, ai_rows)

    # DEPENDENCIAS
    ws = wb["DEPENDENCIAS"]
    clear_data_rows(ws)
    write_rows(ws, build_dependencies(all_bolts))

    # MERGES
    ws = wb["MERGES"]
    clear_data_rows(ws)
    write_rows(ws, MERGES)

    # HISTORIAL
    ws = wb["HISTORIAL"]
    clear_data_rows(ws)
    write_rows(ws, HISTORY)

    # ARTEFACTOS
    ws = wb["ARTEFACTOS"]
    clear_data_rows(ws)
    write_rows(ws, build_artifacts(all_bolts))

    # PMV
    ws = wb["PMV"]
    clear_data_rows(ws)
    write_rows(ws, PMV_ROWS)

    # --- Hojas ampliadas gobierno IA (v1.1) ---
    write_metadata_sheet(wb)

    ensure_sheet(wb, "RESPALDO_Q1", [
        "Dimension", "Art_num", "Autores", "Revista", "DOI", "Clave_cita", "Uso_gobierno_IA",
    ])
    write_rows(wb["RESPALDO_Q1"], RESPALDO_Q1)

    ensure_sheet(wb, "CONTROLES_ISO42001", [
        "ID_Control", "Clausula_42001", "Control", "Evidencia_repo", "Art_num", "DOI", "Estado", "Observacion",
    ])
    write_rows(wb["CONTROLES_ISO42001"], CONTROLES_ISO42001)

    ensure_sheet(wb, "RIESGOS_IA", [
        "ID_Riesgo", "Categoria", "Descripcion", "Probabilidad", "Impacto", "Mitigacion", "Estado", "Evidencia", "Art_num", "DOI",
    ])
    write_rows(wb["RIESGOS_IA"], RIESGOS_IA)

    ensure_sheet(wb, "MODEL_CARD", ["Campo", "Valor", "Fuente"])
    write_rows(wb["MODEL_CARD"], MODEL_CARD)

    ensure_sheet(wb, "VERSIONES_WEB", [
        "Version_web", "Fecha_release", "Nombre_release", "Cambios_principales", "Macroprocesos_afectados",
    ])
    write_rows(wb["VERSIONES_WEB"], VERSIONES_WEB)

    # DASHBOARD fórmulas
    ws = wb["DASHBOARD"]
    ws["A1"] = f"Dashboard Gobierno IA {PLANTILLA_VERSION} — Web {WEB_APP_VERSION} / Plantilla {CONTENT_VERSION}"
    ws["A1"].font = Font(bold=True, size=11, name="Arial")
    ws["B2"] = "=COUNTA(CONTROL_VERSIONES!B:B)-1"
    ws["B3"] = '=COUNTIF(CONTROL_VERSIONES!F:F,"PRODUCCION")+COUNTIF(CONTROL_VERSIONES!F:F,"ACTIVO")'
    ws["B4"] = '=COUNTIF(CONTROL_VERSIONES!F:F,"PRODUCCION")'
    ws["B5"] = '=COUNTIF(PROCESOS!B:B,"Estratégico")'
    ws["B6"] = '=COUNTIF(PROCESOS!B:B,"Misional")'
    ws["B7"] = '=COUNTIF(PROCESOS!B:B,"Apoyo")'
    ws["A9"] = "Controles ISO 42001"
    ws["B9"] = "=COUNTA(CONTROLES_ISO42001!A:A)-1"
    ws["A10"] = "Riesgos IA registrados"
    ws["B10"] = "=COUNTA(RIESGOS_IA!A:A)-1"
    ws["A11"] = "Artículos Q1 respaldo"
    ws["B11"] = "=COUNTA(RESPALDO_Q1!A:A)-1"
    ws["A12"] = "Versión contenido plantilla"
    ws["B12"] = CONTENT_VERSION
    ws["A13"] = "Versión sistema web"
    ws["B13"] = WEB_APP_VERSION
    ws["A14"] = "Fecha generación"
    ws["B14"] = TODAY
    ws["A15"] = "Releases web registrados"
    ws["B15"] = "=COUNTA(VERSIONES_WEB!A:A)-1"

    format_all_sheets(wb)

    OUT_ARTIFACTS.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_ARTIFACTS)
    if OUT_DOWNLOADS.parent.exists():
        shutil.copy2(OUT_ARTIFACTS, OUT_DOWNLOADS)
        print(f"Copia:    {OUT_DOWNLOADS}")

    print(f"Generado: {OUT_ARTIFACTS}")
    print(f"Plantilla: {PLANTILLA_VERSION} | Web: {WEB_APP_VERSION} | Contenido: {CONTENT_VERSION} | Fecha: {TODAY}")
    print(f"Procesos: {len(build_macro_processes())}")
    print(f"Bolts:    {len(all_bolts)}")
    print(f"Requerimientos: {len(REQUIREMENTS)}")
    print(f"AI_DLC:   {len(ai_rows)}")
    print(f"Controles ISO 42001: {len(CONTROLES_ISO42001)}")
    print(f"Riesgos IA: {len(RIESGOS_IA)}")
    print(f"Respaldo Q1: {len(RESPALDO_Q1)}")
    print(f"Releases web: {len(VERSIONES_WEB)}")
    print(f"Artefactos: {len(build_artifacts(all_bolts))}")


if __name__ == "__main__":
    main()
