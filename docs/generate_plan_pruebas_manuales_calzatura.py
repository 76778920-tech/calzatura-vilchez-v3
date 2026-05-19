"""
Genera Plan_Pruebas_Manuales_Calzatura_Vilchez_V3.xlsx:
- Enumera TODOS los tests (Vitest + Playwright + pytest) desde el repo.
- Hoja «Casos automatizados (CI)»: una fila por test, mismas columnas tipo plan Horizonte.
- Hoja «Casos manuales»: plantillas TC manuales.
- Resumen con conteos reales.

Ejecutar desde la raíz del repo:
  python docs/generate_plan_pruebas_manuales_calzatura.py

Requiere: pip install openpyxl
"""
from __future__ import annotations

import re
import shutil
import subprocess
import sys
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
WEB = ROOT / "calzatura-vilchez"
AI = ROOT / "ai-service"
OUT = ROOT / "docs" / "Plan_Pruebas_Manuales_Calzatura_Vilchez_V3.xlsx"

HEADERS = [
    "ID",
    "Módulo",
    "Caso de prueba",
    "Tipo",
    "Prioridad",
    "Pasos / identificación",
    "Resultado esperado",
    "Resultado obtenido",
    "Estado",
    "Evidencia",
    "Archivo / nodo",
    "Herramienta",
]


@dataclass
class TestRow:
    tool: str
    rel_file: str
    module: str
    title: str
    sort_key: tuple[str, ...]


def _node() -> str:
    return shutil.which("node") or "node"


def _npx() -> str:
    return shutil.which("npx") or "npx"


def _run(cmd: list[str], cwd: Path) -> str:
    p = subprocess.run(
        cmd,
        cwd=str(cwd),
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=300,
    )
    out = (p.stdout or "") + ("\n" + (p.stderr or "") if p.stderr else "")
    return out


def module_from_web_path(rel: str) -> str:
    rel = rel.replace("\\", "/")
    if "src/domains/" in rel:
        seg = rel.split("src/domains/", 1)[1].split("/", 1)[0]
        return seg
    if "src/__tests__/" in rel:
        return "__tests__"
    if "src/utils/" in rel or rel.startswith("src/utils/"):
        return "utils"
    return str(Path(rel).parent).replace("\\", "/")


def collect_vitest() -> list[TestRow]:
    out = _run(
        [_node(), str(WEB / "scripts" / "run-clean-env.mjs"), "vitest", "list"],
        WEB,
    )
    rows: list[TestRow] = []
    for line in out.splitlines():
        line = line.strip()
        if not line or line.startswith("WARN") or line.startswith("⎯"):
            continue
        if " > " not in line:
            continue
        parts = [p.strip() for p in line.split(" > ")]
        rel_file = parts[0].replace("\\", "/")
        title = parts[-1]
        suite = " › ".join(parts[1:-1]) if len(parts) > 2 else ""
        full_title = f"{suite} › {title}" if suite else title
        mod = module_from_web_path(rel_file)
        rows.append(
            TestRow(
                tool="Vitest",
                rel_file=rel_file,
                module=mod,
                title=full_title,
                sort_key=("0", rel_file.lower(), full_title.lower()),
            )
        )
    return rows


_RE_PW = re.compile(
    r"^\s*\[[^\]]+\]\s*›\s*(?P<file>[^\s:]+\.spec\.ts):\d+:\d+\s*›\s*(?P<title>.+)$",
)


def collect_playwright() -> list[TestRow]:
    out = _run([_npx(), "--yes", "playwright", "test", "--list"], WEB)
    rows: list[TestRow] = []
    for line in out.splitlines():
        line = line.rstrip()
        if not line or line.startswith("Listing"):
            continue
        m = _RE_PW.match(line)
        if not m:
            continue
        fn = m.group("file")
        title = m.group("title").strip()
        rel = f"e2e/{fn}" if not fn.startswith("e2e/") else fn
        rows.append(
            TestRow(
                tool="Playwright",
                rel_file=rel.replace("\\", "/"),
                module="e2e",
                title=title,
                sort_key=("1", fn.lower(), title.lower()),
            )
        )
    return rows


def collect_pytest() -> list[TestRow]:
    out = _run([sys.executable, "-m", "pytest", "--collect-only", "-q"], AI)
    rows: list[TestRow] = []
    for line in out.splitlines():
        line = line.strip()
        if not line or line.endswith("collected in") or "error" in line.lower() and "===" in line:
            continue
        if "::" not in line or not line.startswith("tests/"):
            continue
        path_part, _, node = line.partition("::")
        rel = path_part.replace("\\", "/")
        mod = Path(rel).stem
        rows.append(
            TestRow(
                tool="pytest",
                rel_file=rel,
                module=mod,
                title=node,
                sort_key=("2", rel.lower(), node.lower()),
            )
        )
    return rows


def autosize(ws, max_col: int, min_w: float = 8, max_w: float = 70) -> None:
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        m = min_w
        for row in range(1, min(ws.max_row, 4000) + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            m = max(m, min(max_w, len(str(v)) * 0.92 + 1.5))
        ws.column_dimensions[letter].width = m


def apply_header_row(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="DDEBF7")
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def sheet_resumen(wb: Workbook, n_v: int, n_e: int, n_p: int, n_man: int) -> None:
    ws = wb.create_sheet("Resumen")
    today = date.today().isoformat()
    total_auto = n_v + n_e + n_p
    ws.append(["INSTRUMENTOS DEL PLAN DE VALIDACIÓN Y ASEGURAMIENTO DE CALIDAD"])
    ws.append(["Calzatura Vilchez V3 — SPA React (Vite), Supabase, Firebase Functions, AI Service (FastAPI), Stripe"])
    ws.merge_cells("A3:E3")
    cell = ws["A3"]
    cell.value = (
        f"IMPORTANTE: cada test automatizado está en UNA fila de la primera pestaña del libro: "
        f"«Casos de prueba (automatizados)» ({total_auto} filas de datos + cabecera). "
        "Ahí verás ID (AUTO-VIT-…, AUTO-E2E-…, AUTO-PY-…), módulo, nombre del caso, archivo y herramienta. "
        "La tabla de abajo en esta hoja solo resume totales (no es el detalle test a test)."
    )
    cell.font = Font(bold=True, size=11, color="006100")
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[3].height = 48
    ws.append(["Instrumento", "Propósito", None, "Indicador", "Valor"])
    data = [
        ("Tests automatizados (CI)", "Detalle: ver primera pestaña «Casos de prueba (automatizados)».", None, "Tests únicos (Vitest + Playwright + pytest)", total_auto),
        ("Vitest", "Regresión TS/React (unit e integración ligera).", None, "Filas con herramienta Vitest en esa hoja", n_v),
        ("Playwright E2E", "Flujos UI (catálogo, admin, carrito, etc.).", None, "Filas con herramienta Playwright en esa hoja", n_e),
        ("pytest (AI Service)", "Modelos, API y clientes.", None, "Filas con herramienta pytest en esa hoja", n_p),
        ("Casos manuales (plantilla)", "Exploración y UAT no automatizados.", None, "Filas en hoja «Casos manuales»", n_man),
        ("Matriz de trazabilidad", "Requerimiento ↔ caso ↔ evidencia.", None, "Hoja «Trazabilidad»", "Ver hoja"),
        ("Checklist release", "Smoke por área.", None, "Hoja «Checklist»", "Ver hoja"),
        ("SonarCloud / CI", "Calidad y pipelines.", None, "Última actualización (generación plan)", today),
    ]
    for r in data:
        ws.append(list(r))
    for c in range(1, 6):
        cell = ws.cell(row=1, column=c)
        if cell.value:
            cell.font = Font(bold=True, size=12)
    autosize(ws, 5)


def sheet_casos_automatizados(wb: Workbook, tests: list[TestRow]) -> None:
    """Primera pestaña del libro: una fila por test."""
    ws = wb.create_sheet("Casos de prueba (automatizados)", 0)
    ws.append(HEADERS)
    apply_header_row(ws, 1)
    tests_sorted = sorted(tests, key=lambda t: t.sort_key)
    prefix = {"Vitest": "AUTO-VIT", "Playwright": "AUTO-E2E", "pytest": "AUTO-PY"}
    counts = {"Vitest": 0, "Playwright": 0, "pytest": 0}
    for tr in tests_sorted:
        counts[tr.tool] = counts[tr.tool] + 1
        n = counts[tr.tool]
        tid = f"{prefix[tr.tool]}-{n:04d}"
        tipo = f"Automático ({tr.tool})"
        pasos = f"Ejecutar suite en CI o local: {tr.tool}. Archivo: {tr.rel_file}"
        ws.append(
            [
                tid,
                tr.module,
                tr.title,
                tipo,
                "Media",
                pasos,
                "PASS (sin regresión)",
                "",
                "Automatizado (CI)",
                "GitHub Actions / reporte HTML del runner",
                tr.rel_file,
                tr.tool,
            ]
        )
    ws.freeze_panes = "A2"
    last_row = ws.max_row
    last_col = len(HEADERS)
    tab = Table(
        displayName="TablaCasosPruebaAuto",
        ref=f"A1:{get_column_letter(last_col)}{last_row}",
    )
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)
    autosize(ws, last_col)


def sheet_casos_manuales(wb: Workbook) -> None:
    ws = wb.create_sheet("Casos manuales")
    ws.append(HEADERS)
    apply_header_row(ws, 1)
    cases: list[tuple] = [
        ("TC-M001", "Público / catálogo", "Listado y filtros del catálogo", "Manual / UAT", "Alta", "Recorrido usuario; comparar con datos Supabase.", "Coherencia visual y datos.", "", "PENDIENTE", "", "catálogo público", "Manual"),
        ("TC-M002", "PDP", "Ficha producto: tallas, precio, imágenes", "Manual", "Alta", "Abrir PDP de producto activo y oculto.", "Reglas de visibilidad correctas.", "", "PENDIENTE", "", "PDP", "Manual"),
        ("TC-M003", "Carrito", "Límites de stock y mensajes", "Manual", "Alta", "Forzar cantidades límite.", "Mensajes claros.", "", "PENDIENTE", "", "carrito", "Manual"),
        ("TC-M004", "Checkout Stripe", "Flujo pago cancelación / éxito staging", "Manual", "Alta", "Tarjeta test Stripe.", "Redirecciones correctas.", "", "PENDIENTE", "", "Stripe", "Manual"),
        ("TC-M005", "Admin ventas", "Registro venta diaria multi-línea", "Manual", "Alta", "Operación real en staging.", "Stock y totales coherentes.", "", "PENDIENTE", "", "RPC ventas", "Manual"),
        ("TC-M006", "Admin productos", "Variantes e imágenes", "Manual", "Alta", "CRUD completo en staging.", "Persistencia OK.", "", "PENDIENTE", "", "productos", "Manual"),
        ("TC-M007", "RLS Supabase", "Aislamiento entre usuarios", "Manual / SQL", "Alta", "Dos sesiones.", "Sin fugas de datos.", "", "PENDIENTE", "", "RLS", "SQL + app"),
        ("TC-M008", "AI Service", "Contrato API con token", "Manual", "Media", "curl / Postman.", "401/200 según token.", "", "PENDIENTE", "", "FastAPI", "Manual"),
    ]
    for row in cases:
        ws.append(list(row))
    ws.freeze_panes = "A2"
    autosize(ws, len(HEADERS))


def sheet_checklist(wb: Workbook) -> None:
    ws = wb.create_sheet("Checklist")
    ws.append(["ID", "Área", "Elemento validado", "Criterio de validación", "Estado", "Observaciones", "Evidencia", "Responsable"])
    for i, (area, el, crit) in enumerate(
        [
            ("Frontend", "Auth cliente", "Login / sesión Firebase"),
            ("Frontend", "Catálogo + PDP", "Productos y precios coherentes"),
            ("Frontend", "Carrito + checkout", "Totales y Stripe"),
            ("Supabase", "RLS", "Datos aislados por usuario/rol"),
            ("Functions", "Pedidos / webhook", "Errores controlados"),
            ("AI Service", "Proxy admin", "Bearer y upstream"),
            ("CI", "GitHub Actions", "Jobs en verde"),
            ("Calidad", "SonarCloud", "Gate acordado"),
        ],
        start=1,
    ):
        ws.append([f"CHK-{i:03d}", area, el, crit, "PENDIENTE", "", f"evidencias/CHK-{i:03d}", "QA"])
    ws.freeze_panes = "A2"
    autosize(ws, 8)


def sheet_defectos(wb: Workbook) -> None:
    ws = wb.create_sheet("Registro defectos")
    ws.append(
        [
            "ID Defecto",
            "Módulo",
            "Archivo / Servicio",
            "Descripción",
            "Severidad",
            "Prioridad",
            "Estado",
            "Evidencia",
            "Corrección",
            "Responsable",
            "Acción correctiva",
            "Fecha cierre",
        ]
    )
    apply_header_row(ws, 1)
    ws.append(["(Plantilla)", "-", "-", "Registrar hallazgos manuales o de exploración.", "-", "-", "Abierto", "", "", "", "", ""])
    ws.freeze_panes = "A2"
    autosize(ws, 12)


def sheet_trazabilidad(wb: Workbook, n_auto: int) -> None:
    ws = wb.create_sheet("Trazabilidad")
    ws.append(["ID", "Requerimiento / ámbito", "Nota", "Módulo", "Caso vinculado", "Herramienta", "Evidencia", "Estado"])
    ws.append(["TRZ-AUTO", "Suite automatizada", f"{n_auto} tests en hoja «Casos de prueba (automatizados)»", "Repo", "AUTO-*", "Vitest / Playwright / pytest", "CI artifacts", "Vivo (CI)"])
    ws.append(["TRZ-MAN", "UAT manual", "Hoja «Casos manuales»", "Negocio", "TC-M*", "Manual", "Capturas", "PENDIENTE"])
    ws.freeze_panes = "A2"
    autosize(ws, 8)


def sheet_criterios(wb: Workbook) -> None:
    ws = wb.create_sheet("Criterios aceptación")
    ws.append(["ID", "Módulo", "Funcionalidad", "Criterio de aceptación", "Validación", "Resultado esperado", "Estado"])
    rows = [
        ("CA-001", "SPA", "Catálogo", "Solo productos conforme reglas de visibilidad.", "Vitest + E2E + manual", "Listado coherente", "PENDIENTE"),
        ("CA-002", "SPA", "Carrito / stock", "No vender por encima del stock por variante.", "Vitest stock + E2E", "Bloqueo o tope", "PENDIENTE"),
        ("CA-003", "Supabase", "RPC admin", "Operaciones atómicas sin estados rotos.", "E2E admin + RPC", "Datos coherentes", "PENDIENTE"),
        ("CA-004", "AI Service", "API", "Contratos y auth documentados.", "pytest", "Respuestas válidas", "PENDIENTE"),
    ]
    for r in rows:
        ws.append(list(r))
    ws.freeze_panes = "A2"
    autosize(ws, 7)


def sheet_catalogos(wb: Workbook) -> None:
    ws = wb.create_sheet("Catalogos")
    ws.append(["CATÁLOGO DE MÓDULOS — Calzatura Vilchez V3"])
    ws.append([])
    ws.append(["Código", "Módulo", "Descripción"])
    mods = [
        ("MOD-01", "SPA React (Vite)", "Tienda, carrito, checkout, cuenta."),
        ("MOD-02", "Supabase", "Postgres, RLS, RPC."),
        ("MOD-03", "Firebase", "Auth cliente + Functions."),
        ("MOD-04", "AI Service", "FastAPI predicción / campañas."),
        ("MOD-05", "Stripe", "Pagos."),
        ("MOD-06", "CI / Sonar", "GitHub Actions."),
    ]
    for m in mods:
        ws.append(list(m))
    style_title = ws.cell(row=1, column=1)
    style_title.font = Font(bold=True, size=12)
    autosize(ws, 3)


def sheet_correcciones(wb: Workbook) -> None:
    ws = wb.create_sheet("Correcciones y ampliaciones")
    ws.append(["ID", "Área", "Observación / corrección", "Estado anterior", "Corrección aplicada", "Impacto", "Estado final"])
    ws.append(["CORR-000", "Documentación QA", "Plan Excel generado desde enumeración real de tests.", "-", "Script docs/generate_plan_pruebas_manuales_calzatura.py", "Trazabilidad test↔fila", "Hecho"])
    ws.freeze_panes = "A2"
    autosize(ws, 7)


def main() -> None:
    print("Recolectando Vitest…", flush=True)
    vit = collect_vitest()
    print(f"  Vitest: {len(vit)}", flush=True)
    print("Recolectando Playwright…", flush=True)
    pw = collect_playwright()
    print(f"  Playwright: {len(pw)}", flush=True)
    print("Recolectando pytest…", flush=True)
    py = collect_pytest()
    print(f"  pytest: {len(py)}", flush=True)
    all_auto = vit + pw + py
    n_man = 8

    wb = Workbook()
    d = wb.active
    wb.remove(d)

    sheet_casos_automatizados(wb, all_auto)
    sheet_resumen(wb, len(vit), len(pw), len(py), n_man)
    sheet_casos_manuales(wb)
    sheet_checklist(wb)
    sheet_defectos(wb)
    sheet_trazabilidad(wb, len(all_auto))
    sheet_criterios(wb)
    sheet_catalogos(wb)
    sheet_correcciones(wb)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"\nEscrito: {OUT}", flush=True)
    print(f"Total automatizados: {len(all_auto)} (= {len(vit)} + {len(pw)} + {len(py)})", flush=True)


if __name__ == "__main__":
    main()
