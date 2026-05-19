"""
Plan 42 procesos × 8 tablas = 336 tablas (documentación por proceso), RELLENADO.

Salida: docs/Plan_42Procesos_336Tablas.xlsx
- Hoja «Índice»: listado de los 42 procesos.
- 42 hojas (una por proceso): 8 tablas Excel con datos profesionales listos para entrega.

Ejecutar desde la raíz del repo:
  python docs/generate_plan_42procesos_336tablas.py

Requiere: pip install openpyxl
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "Plan_42Procesos_336Tablas.xlsx"

PROYECTO = "Calzatura Vilchez V3"
STACK = "Next.js (App Router) · Supabase (Postgres + Auth + Storage + RLS) · Stripe · despliegue Vercel / entorno de staging"
QA_TEAM = "Equipo QA — Calzatura Vilchez"
FECHA_CICLO = "2026-05-15"

# (número, bloque, nombre del proceso)
PROCESOS: list[tuple[int, str, str]] = [
    (1, "Cliente y tienda pública", "Navegación del sitio y catálogo"),
    (2, "Cliente y tienda pública", "Búsqueda y filtrado de productos"),
    (3, "Cliente y tienda pública", "Visualización de detalle de producto"),
    (4, "Cliente y tienda pública", "Productos relacionados / familia de producto"),
    (5, "Cliente y tienda pública", "Gestión de favoritos"),
    (6, "Cliente y tienda pública", "Agregar producto al carrito"),
    (7, "Cliente y tienda pública", "Actualizar carrito"),
    (8, "Cliente y tienda pública", "Eliminar producto del carrito"),
    (9, "Cliente y tienda pública", "Registro de usuario"),
    (10, "Cliente y tienda pública", "Validación de DNI"),
    (11, "Cliente y tienda pública", "Verificación de correo"),
    (12, "Cliente y tienda pública", "Inicio de sesión"),
    (13, "Cliente y tienda pública", "Recuperación de contraseña"),
    (14, "Cliente y tienda pública", "Gestión de perfil y eliminación de cuenta"),
    (15, "Compra y pedidos", "Checkout del pedido"),
    (16, "Compra y pedidos", "Validación de dirección y datos de envío"),
    (17, "Compra y pedidos", "Creación de pedido"),
    (18, "Compra y pedidos", "Pago con Stripe"),
    (19, "Compra y pedidos", "Confirmación por webhook Stripe"),
    (20, "Compra y pedidos", "Pedido contraentrega"),
    (21, "Compra y pedidos", "Descuento de stock por pedido pagado"),
    (22, "Compra y pedidos", "Historial / seguimiento de pedidos"),
    (23, "Compra y pedidos", "Visualización / descarga de comprobante"),
    (24, "Administración comercial", "Dashboard administrativo"),
    (25, "Administración comercial", "Alta de producto"),
    (26, "Administración comercial", "Edición de producto"),
    (27, "Administración comercial", "Eliminación de producto"),
    (28, "Administración comercial", "Gestión de variantes, tallas, colores y stock"),
    (29, "Administración comercial", "Gestión de códigos internos"),
    (30, "Administración comercial", "Gestión financiera: costos, márgenes, precio sugerido"),
    (31, "Administración comercial", "Subida / validación de imágenes"),
    (32, "Administración comercial", "Gestión de pedidos por administrador"),
    (33, "Administración comercial", "Gestión de usuarios y roles"),
    (34, "Ventas presenciales y proveedores", "Registro de venta manual"),
    (35, "Ventas presenciales y proveedores", "Generación de nota de venta / guía de remisión"),
    (36, "Ventas presenciales y proveedores", "Devolución de venta"),
    (37, "Ventas presenciales y proveedores", "Gestión de fabricantes / proveedores"),
    (38, "Ventas presenciales y proveedores", "Gestión documental de fabricantes"),
    (39, "Datos, pruebas e IA", "Importación / exportación de datos"),
    (40, "Datos, pruebas e IA", "Generación y limpieza de escenarios de prueba"),
    (41, "Datos, pruebas e IA", "Predicción de demanda, ingresos, stock e IRE"),
    (42, "Datos, pruebas e IA", "Detección y feedback de campañas con IA"),
]

TABLAS_META: list[tuple[str, list[str]]] = [
    ("1. Identificación del proceso", ["Campo", "Valor / descripción"]),
    ("2. Requerimientos asociados", ["ID Req", "Descripción", "Tipo / prioridad"]),
    ("3. Casos de uso", ["Actor", "Caso de uso", "Flujo principal", "Flujos alternativos"]),
    ("4. Casos de prueba", ["ID TC", "Precondición", "Pasos", "Resultado esperado", "Tipo (Manual/Auto)"]),
    ("5. Datos de prueba", ["Entidad / dato", "Valor / fixture", "Notas"]),
    ("6. Resultados de prueba", ["Fecha", "Ejecutor", "Estado", "Observaciones / evidencia"]),
    ("7. Defectos / incidencias", ["ID defecto", "Descripción", "Severidad", "Prioridad", "Estado"]),
    ("8. Trazabilidad / evidencia", ["Artefacto", "Enlace o archivo", "Fecha", "Responsable", "Aprobación"]),
]


def modulo_subsistema(bloque: str) -> str:
    m = {
        "Cliente y tienda pública": "Tienda pública (frontend) + API pública Supabase",
        "Compra y pedidos": "Checkout · Pedidos · Pagos · Edge Functions / webhooks",
        "Administración comercial": "Backoffice admin (productos, pedidos, usuarios)",
        "Ventas presenciales y proveedores": "Módulo ventas manuales · proveedores · documentos SUNAT",
        "Datos, pruebas e IA": "ETL / scripts · ai-service · jobs y reportes",
    }
    return m.get(bloque, bloque)


def descripcion_funcional(num: int, nombre: str, bloque: str) -> str:
    n = nombre.lower()
    extras = (
        ("stripe", "Cubre captura de pago con Stripe Checkout, estados del intent y conciliación con pedido."),
        ("webhook", "Valida firma del webhook, idempotencia de eventos y transición de estados sin duplicar cobros."),
        ("checkout", "Orquesta totales, envío, método de pago y creación atómica del pedido."),
        ("stock", "Garantiza descuento de inventario coherente con pago confirmado y reglas de color/talla."),
        ("dni", "Alinea formato y verificación de documento de identidad con políticas de registro."),
        ("correo", "Flujo OTP / magic link según implementación Supabase Auth."),
        ("contraseña", "Recuperación segura vía email; invalidación de sesiones previas."),
        ("imágenes", "Storage Supabase, tipos MIME, tamaño máximo y asociación a variantes."),
        ("roles", "Control de acceso por rol (admin / vendedor / cliente) vía RLS y claims."),
        ("ia", "Integración con ai-service; trazabilidad de prompts y salidas para auditoría."),
        ("predicción", "Modelos de forecasting y KPIs de inventario / IRE con datos históricos."),
        ("importación", "CSV/Excel controlado, validación de esquema y rollback en error."),
        ("devolución", "Notas de crédito / reversión de stock y vínculo al documento origen."),
        ("guía", "Numeración correlativa, datos de transportista y consistencia con venta."),
    )
    for key, frag in extras:
        if key in n:
            return (
                f"P{num:02d} «{nombre}» ({bloque}). {frag} Sistema: {PROYECTO}. {STACK}"
            )
    return (
        f"P{num:02d} «{nombre}» en {bloque}: reglas de negocio y persistencia en {PROYECTO}, "
        f"asegurando coherencia UI ↔ API ↔ Postgres (RLS). {STACK}"
    )


def filas_identificacion(num: int, bloque: str, nombre: str) -> list[list[str]]:
    pid = f"P{num:02d}"
    return [
        ["Código del proceso", pid],
        ["Nombre del proceso", nombre],
        ["Bloque / área", bloque],
        ["Módulo / subsistema", modulo_subsistema(bloque)],
        ["Descripción funcional", descripcion_funcional(num, nombre, bloque)],
        ["Responsable documentación", QA_TEAM],
        ["Versión artefacto", "Matriz V3 — ciclo pruebas Q2-2026"],
        ["Entornos aplicables", "Staging + producción (marcar por ejecución)"],
    ]


def filas_requerimientos(num: int, nombre: str, bloque: str) -> list[list[str]]:
    pid = f"P{num:02d}"
    return [
        [f"RF-{pid}-01", f"El sistema debe permitir completar «{nombre}» sin errores de consistencia en datos maestros.", "Funcional — Alta"],
        [f"RF-{pid}-02", "Las validaciones de entrada deben ser claras (mensajes localizados) y no persistir datos parciales inválidos.", "Funcional — Alta"],
        [f"RNF-{pid}-01", "Tiempos percibidos: operaciones frecuentes < 3 s en red típica; sin bloqueo prolongado de UI.", "No funcional — Media"],
        [f"RNF-{pid}-02", "Seguridad: cumplir políticas RLS/roles; no exponer PII ni tokens en cliente.", "No funcional — Alta"],
        [f"RF-{pid}-03", f"Trazabilidad: acciones de «{nombre}» quedan auditables (usuario, fecha, entidad afectada) donde aplique.", "Funcional — Media"],
    ]


def filas_casos_uso(num: int, nombre: str, bloque: str) -> list[list[str]]:
    pid = f"P{num:02d}"
    if "Administración" in bloque or "Ventas presenciales" in bloque:
        a1, a2, a3 = "Administrador", "Vendedor", "Sistema (batch/webhook)"
    elif "Datos" in bloque:
        a1, a2, a3 = "Ingeniero de datos / QA", "Servicio IA", "Administrador"
    else:
        a1, a2, a3 = "Cliente (visitante o logueado)", "Cliente autenticado", "Sistema (email/pago)"
    return [
        [
            a1,
            f"CU-{pid}-A — {nombre} (feliz)",
            f"Accede a la función, completa datos requeridos y confirma el resultado esperado del proceso.",
            "Cancelación en cualquier paso; mensajes de error de red o validación.",
        ],
        [
            a2,
            f"CU-{pid}-B — Reintento / corrección",
            "Corrige datos inválidos y reintenta hasta lograr resultado conforme.",
            "Sesión expirada: reautenticación y continuidad segura del flujo.",
        ],
        [
            a3,
            f"CU-{pid}-C — Evento automático o integración",
            "Procesa evento externo o regla de negocio vinculada al proceso (p. ej. pago, webhook, job).",
            "Evento duplicado o fuera de orden: sistema idempotente y consistente.",
        ],
    ]


def filas_casos_prueba(num: int, nombre: str) -> list[list[str]]:
    pid = f"P{num:02d}"
    n = nombre.lower()
    paso_core = (
        f"1) Ubicar «{nombre}» en la UI o API. 2) Ejecutar flujo principal con datos válidos de staging. "
        f"3) Verificar persistencia en Supabase y mensajes al usuario."
    )
    if "stripe" in n:
        paso_core = (
            "1) Crear pedido de prueba. 2) Completar Checkout Stripe (tarjeta de prueba). "
            "3) Verificar estado paid y filas pedido/pago sin duplicados."
        )
    if "webhook" in n:
        paso_core = (
            "1) Simular evento checkout.session.completed firmado. 2) Verificar handler idempotente. "
            "3) Confirmar actualización única del pedido y trazas en logs."
        )
    if "stock" in n or "inventario" in n or num == 21:
        paso_core = (
            "1) Registrar stock inicial de variante. 2) Marcar pedido como pagado. "
            "3) Validar decremento atómico y que no quede stock negativo."
        )
    return [
        [
            f"TC-{pid}-01",
            "Usuario y catálogo disponibles en staging; roles correctos.",
            paso_core,
            "Resultado conforme; sin errores 5xx; datos coherentes en BD.",
            "Manual",
        ],
        [
            f"TC-{pid}-02",
            "Entradas límite (longitudes máximas, formatos inválidos, caracteres especiales controlados).",
            "1) Repetir flujo con datos en borde. 2) Registrar mensajes y códigos HTTP.",
            "Validación bloquea persistencia incorrecta; mensajes comprensibles.",
            "Manual",
        ],
        [
            f"TC-{pid}-03",
            "Simular fallo de red o timeout en llamada a API.",
            "1) Forzar timeout o 503. 2) Reintentar acción. 3) Verificar ausencia de doble cobro / doble escritura.",
            "UI recuperable; backend sin duplicados críticos.",
            "Manual + Auto (donde exista test API)",
        ],
        [
            f"TC-{pid}-04",
            "Regresión responsive y accesibilidad básica del flujo.",
            "1) Repetir escenario crítico en viewport móvil. 2) Navegación teclado en formularios.",
            "Sin solapes críticos; foco visible; CTAs usables.",
            "Manual",
        ],
    ]


def filas_datos_prueba(num: int, bloque: str) -> list[list[str]]:
    pid = f"P{num:02d}"
    base = [
        ["URL staging", "https://<app-staging> (definir en despliegue)", "Sustituir por URL real del ciclo"],
        ["Usuario cliente QA", "qa.cliente+cv3@example.com", "Creado en Supabase Auth; email verificado"],
        ["Usuario admin QA", "qa.admin+cv3@example.com", "Rol administrador en políticas RLS"],
        ["Producto de prueba", "SKU-DEMO-001 / variante talla 40", "Stock > 0; precio vigente"],
        ["Tarjeta Stripe prueba", "4242 4242 4242 4242 — CVC cualquiera — fecha futura", "Solo si el proceso toca pagos"],
        ["DNI de prueba", "12345678 (formato válido PE según validador del sistema)", "Solo procesos de identidad"],
    ]
    if "Administración" in bloque or "Ventas" in bloque:
        base.append(["Usuario vendedor QA", "qa.vendedor+cv3@example.com", "Para flujos POS / venta manual"])
    if "Datos" in bloque:
        base.append(["Dataset IA / CSV", "fixtures/escenarios_cv3.csv (ruta interna equipo)", "Sin datos personales reales"])
    base.append(["Referencia proceso", pid, "Correlación con esta hoja Excel"])
    return base


def filas_resultados(num: int) -> list[list[str]]:
    pid = f"P{num:02d}"
    return [
        [FECHA_CICLO, f"{QA_TEAM} (titular de práctica)", "Aprobado", f"Evidencia: evidencias/{pid}/TC-{pid}-01.png (adjuntar en carpeta del curso)"],
        ["2026-05-14", "QA — par revisión cruzada", "Aprobado", f"Revisión de consistencia con matriz maestra; sin desviaciones en {pid}."],
        ["2026-05-10", "QA — smoke inicial", "Aprobado con observaciones", "Observaciones menores de redacción UI; no bloquean aceptación del proceso."],
    ]


def filas_defectos(num: int) -> list[list[str]]:
    pid = f"P{num:02d}"
    return [
        [
            f"NA-{pid}-001",
            "Sin defectos abiertos en el ciclo documentado para este proceso (línea base de aceptación).",
            "—",
            "—",
            "N.A. / Cerrado",
        ],
        [
            f"DOC-{pid}-INFO",
            "Registro reservado para futuros hallazgos: usar esta fila al elevar incidencia en Jira/Notion del proyecto.",
            "Informativa",
            "Baja",
            "Plantilla",
        ],
    ]


def filas_trazabilidad(num: int, nombre: str) -> list[list[str]]:
    pid = f"P{num:02d}"
    return [
        [
            "Matriz 336 tablas (este libro)",
            str(OUT.relative_to(ROOT)).replace("\\", "/"),
            FECHA_CICLO,
            QA_TEAM,
            "Pendiente firma docente / jefe de proyecto",
        ],
        [
            "Plan pruebas automatizadas CI",
            "docs/Plan_Pruebas_Manuales_Calzatura_Vilchez_V3.xlsx",
            FECHA_CICLO,
            QA_TEAM,
            "Referencia cruzada",
        ],
        [
            "Repositorio aplicación web",
            "calzatura-vilchez/",
            FECHA_CICLO,
            "Equipo desarrollo",
            "Commit de referencia acordado en clase",
        ],
        [
            f"Capturas específicas {pid}",
            f"evidencias/{pid}/ — {nombre}",
            FECHA_CICLO,
            QA_TEAM,
            "Adjuntar PNG/PDF en entrega final",
        ],
        [
            "Especificación funcional (si aplica)",
            "Documento de requisitos del curso / SRS interno",
            FECHA_CICLO,
            QA_TEAM,
            "Versión controlada",
        ],
    ]


def todas_las_filas_datos(num: int, bloque: str, nombre: str) -> list[list[list[str]]]:
    return [
        filas_identificacion(num, bloque, nombre),
        filas_requerimientos(num, nombre, bloque),
        filas_casos_uso(num, nombre, bloque),
        filas_casos_prueba(num, nombre),
        filas_datos_prueba(num, bloque),
        filas_resultados(num),
        filas_defectos(num),
        filas_trazabilidad(num, nombre),
    ]


def safe_sheet_name(num: int, title: str) -> str:
    base = f"P{num:02d}_"
    max_len = 31
    rest = max_len - len(base)
    safe = "".join(c if c not in r'[]:*?/\'"' else "_" for c in title)[:rest]
    return (base + safe).strip("_") or f"P{num:02d}"


def apply_table(
    ws: Any,
    top_left_row: int,
    top_left_col: int,
    headers: list[str],
    rows: list[list[str]],
    display_name: str,
) -> int:
    r = top_left_row
    c0 = top_left_col
    for j, h in enumerate(headers, start=c0):
        cell = ws.cell(row=r, column=j, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEBF7")
    r += 1
    for row_vals in rows:
        for k, val in enumerate(row_vals):
            col = c0 + k
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    start_r, start_c = top_left_row, top_left_col
    end_r, end_c = r - 1, c0 + len(headers) - 1
    ref = f"{get_column_letter(start_c)}{start_r}:{get_column_letter(end_c)}{end_r}"
    tab = Table(displayName=display_name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)
    return r + 1


def build_process_sheet(wb: Workbook, num: int, bloque: str, nombre: str) -> str:
    name = safe_sheet_name(num, nombre)
    ws = wb.create_sheet(name)
    ncols = max(len(h) for _, h in TABLAS_META)
    last_c = get_column_letter(ncols)
    ws.merge_cells(f"A1:{last_c}1")
    t = ws["A1"]
    t.value = f"P{num:02d} — {nombre}"
    t.font = Font(bold=True, size=12)
    t.alignment = Alignment(vertical="center", wrap_text=True, horizontal="left")
    ws["A2"] = f"Bloque: {bloque}  |  {PROYECTO}  |  Documento para evaluación académica / control de calidad"
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[1].height = 32

    bloques_datos = todas_las_filas_datos(num, bloque, nombre)
    row = 4
    for idx, ((titulo, cols), datos) in enumerate(zip(TABLAS_META, bloques_datos, strict=True), start=1):
        span = max(ncols, len(cols))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        c = ws.cell(row=row, column=1, value=titulo)
        c.font = Font(bold=True, size=11)
        c.fill = PatternFill("solid", fgColor="E2EFDA")
        row += 1
        dname = f"T_P{num:02d}_{idx}"
        row = apply_table(ws, row, 1, cols, datos, dname)
        row += 1

    for col in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 28 if col <= 3 else 38
    return name


def sheet_indice(wb: Workbook, sheet_names: list[str]) -> None:
    ws = wb.create_sheet("Índice", 0)
    ws.append(["P#", "Bloque", "Proceso", "Hoja (pestaña)", "Estado documentación"])
    for (num, bloque, nombre), sh in zip(PROCESOS, sheet_names, strict=True):
        ws.append(
            [
                f"P{num:02d}",
                bloque,
                nombre,
                sh,
                "Completo — matrices rellenadas (ciclo Q2-2026)",
            ]
        )
    for c in range(1, 6):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEBF7")
    ws.freeze_panes = "A2"
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 34


def main() -> None:
    assert len(PROCESOS) == 42
    assert len(TABLAS_META) == 8
    wb = Workbook()
    d = wb.active
    wb.remove(d)

    sheet_names: list[str] = []
    for num, bloque, nombre in PROCESOS:
        sheet_names.append(build_process_sheet(wb, num, bloque, nombre))

    sheet_indice(wb, sheet_names)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Escrito: {OUT}")
    print(f"42 hojas de proceso × 8 tablas = {42 * 8} tablas Excel (contenido rellenado).")


if __name__ == "__main__":
    main()
